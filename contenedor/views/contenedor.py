from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from contenedor.models import Contenedor
from contenedor.serializers.contenedor import ContenedorSerializador, ContenedorActualizarSerializador
from contenedor.serializers.usuario_contenedor import UsuarioContenedorSerializador
from general.serializers.empresa import GenEmpresaSerializador
from general.serializers.configuracion import GenConfiguracionSerializador
from contenedor.models import User
from django.core.management import call_command
from django.shortcuts import get_object_or_404
from django.db import connection
from django.db.models import Sum, Count, Q
from decouple import config
from utilidades.space_do import SpaceDo
from django_tenants.utils import schema_context
import os
import io
from datetime import datetime
from django.utils import timezone
from django.http import HttpResponse
from threading import Thread

def cargar_fixtures_en_segundo_plano(schema_name):
    """
    Función que se ejecutará en segundo plano para cargar los fixtures
    """
    try:
        with schema_context(schema_name):
            # Opción 1: Usando call_command (recomendado)
            #call_command('loaddata', 'fixture1.json', verbosity=0)
            #call_command('loaddata', 'fixture2.json', verbosity=0)
            
            # Opción 2: Manteniendo tu enfoque actual con os.system
            os.system(f"python manage.py tenant_command actualizar_fixtures general/fixtures/ --schema={schema_name}")
            os.system(f"python manage.py tenant_command actualizar_fixtures general/fixtures_inicio/ --schema={schema_name}")                
        print(f"Fixtures cargados exitosamente para {schema_name}")
    except Exception as e:
        print(f"Error cargando fixtures para {schema_name}: {str(e)}")

class ContenedorViewSet(viewsets.ModelViewSet):
    queryset = Contenedor.objects.all()
    serializer_class = ContenedorSerializador    
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self, pk):
        return get_object_or_404(Contenedor, pk=pk)

    def create(self, request):
        try:            
            subdominio = request.data.get('subdominio')
            usuario_id = request.data.get('usuario_id')            
            nombre = request.data.get('nombre')          
            telefono = request.data.get('telefono')
            correo = request.data.get('correo')
            if subdominio and usuario_id and nombre and telefono and correo:
                contenedorValidacion = Contenedor.objects.filter(**{'schema_name':subdominio})
                if contenedorValidacion:
                    return Response({'mensaje': f"Ya existe una empresa con el nombre {subdominio}", "codigo": 13}, status=status.HTTP_400_BAD_REQUEST)
                dominio = '.' + config('DOMINIO_BACKEND')
                usuario = User.objects.get(pk=usuario_id)
                imagenReferencia = f"escandio/logo_defecto.jpg"
                call_command('create_tenant',
                             schema_name=subdominio,
                             domain_domain=subdominio+dominio,
                             nombre=nombre,
                             domain_is_primary='0',
                             imagen=imagenReferencia,
                             usuarios=1,
                             usuario_id=usuario.id,
                             interactive=False)
                #os.system(f"python manage.py tenant_command actualizar_fixtures general/fixtures/ --schema={subdominio}")
                #os.system(f"python manage.py tenant_command actualizar_fixtures general/fixtures_inicio/ --schema={subdominio}")                                           
                thread = Thread(
                    target=cargar_fixtures_en_segundo_plano,
                    args=(subdominio,),
                    daemon=True
                )
                thread.start()                
                
                contenedor = Contenedor.objects.filter(**{'schema_name':subdominio}).first()                        
                data = {'usuario': usuario.id, 'contenedor': contenedor.id, 'rol': 'propietario'}
                usuarioContenedorSerializador = UsuarioContenedorSerializador(data=data)            
                if usuarioContenedorSerializador.is_valid():
                    usuarioContenedorSerializador.save()
                    with schema_context(subdominio):
                        data = {
                            'id':1,
                            'nombre_corto': nombre,
                            'telefono': telefono,
                            'correo': correo,
                            'imagen': imagenReferencia,
                            'contenedor_id':contenedor.id,
                            'subdominio':subdominio}
                        empresaSerializador = GenEmpresaSerializador(data=data)                        
                        if empresaSerializador.is_valid():
                            empresaSerializador.save()
                            data = {
                                'id':1,
                                'empresa':1,
                                'formato_factura':'F'}
                            configuracionSerializador = GenConfiguracionSerializador(data=data)                                                
                            if configuracionSerializador.is_valid():
                                configuracionSerializador.save()                            
                                return Response({'contenedor': usuarioContenedorSerializador.data}, status=status.HTTP_200_OK)            
                            return Response({'mensaje':'Errores en la creacion de la econfiguracion', 'codigo':12, 'validaciones': configuracionSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)
                        return Response({'mensaje':'Errores en la creacion de la empresa', 'codigo':12, 'validaciones': empresaSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)
                return Response({'mensaje':'Errores en la creacion del contenedor', 'codigo':12, 'validaciones': usuarioContenedorSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)
            return Response({'mensaje': 'Faltan datos para el consumo de la api', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)                   
        except User.DoesNotExist:
            return Response({'mensaje':'No existe el usuario para crear la empresa', 'codigo':17}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'mensaje': f'Error interno: {str(e)}', 'codigo': 500}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, pk=None):
        contenedor = self.get_object(pk)
        contenedorSerializador = self.serializer_class(contenedor)
        return Response(contenedorSerializador.data)

    def update(self, request, pk=None):
        empresa = self.get_object(pk)
        empresaSerializador = ContenedorActualizarSerializador(empresa, data=request.data)
        if empresaSerializador.is_valid():
            empresaSerializador.save()
            return Response({'actualizacion': True, 'empresa': empresaSerializador.data}, status=status.HTTP_201_CREATED)            
        return Response({'mensaje':'Errores en la actualizacion', 'codigo':23, 'validaciones': empresaSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None, *args, **kwargs):
        empresa = self.get_object(pk)        
        self.perform_destroy(empresa)
        return Response(status=status.HTTP_200_OK)
       
    @action(detail=False, methods=["post"], url_path=r'validar',)
    def validar(self, request):
        try:
            subdominio = request.data.get('subdominio')
            Contenedor.objects.get(schema_name=subdominio)
            return Response({'mensaje': f"Ya existe una empresa con el nombre {subdominio}", "codigo": 13}, status=status.HTTP_400_BAD_REQUEST)    
        except Contenedor.DoesNotExist:
            return Response({'validar':True}, status=status.HTTP_200_OK)        
        
    @action(detail=False, methods=["post"], url_path=r'cargar-logo',)
    def cargar_logo(self, request):
        try:
            raw = request.data
            empresa_id = raw.get('empresa_id')
            imagenB64 = raw.get('imagenB64')
            if empresa_id:
                empresa = Contenedor.objects.get(pk=empresa_id)
                arrDatosB64 = imagenB64.split(",")
                base64Crudo = arrDatosB64[1]
                arrTipo = arrDatosB64[0].split(";")
                arrData = arrTipo[0].split(":")
                contentType = arrData[1]
                archivo = f"escandio/{config('ENV')}/contenedor/logo_{empresa_id}.jpg"
                spaceDo = SpaceDo()
                spaceDo.putB64(archivo, base64Crudo, contentType)
                empresa.imagen = archivo
                empresa.save()
                return Response({'cargar':True, 'imagen':f"https://{config('DO_BUCKET')}.{config('DO_REGION')}.digitaloceanspaces.com/{archivo}"}, status=status.HTTP_200_OK)                  
            else: 
                return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
        except Contenedor.DoesNotExist:
            return Response({'mensaje':'La empresa no existe', 'codigo':15}, status=status.HTTP_404_NOT_FOUND)  

    @action(detail=False, methods=["post"], url_path=r'limpiar-logo',)
    def limpiar_logo(self, request):
        try:
            raw = request.data
            empresa_id = raw.get('empresa_id')    
            if empresa_id:
                empresa = Contenedor.objects.get(pk=empresa_id)                
                spaceDo = SpaceDo()
                spaceDo.eliminar(empresa.imagen)
                empresa.imagen = f"escandio/logo_defecto.jpg"
                empresa.save()
                return Response({'limpiar':True, 'imagen':f"https://{config('DO_BUCKET')}.{config('DO_REGION')}.digitaloceanspaces.com/{empresa.imagen}"}, status=status.HTTP_200_OK)                  
            else: 
                return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
        except Contenedor.DoesNotExist:
            return Response({'mensaje':'La empresa no existe', 'codigo':15}, status=status.HTTP_404_NOT_FOUND)     

    @action(detail=False, methods=["post"], url_path=r'toggle-whatsapp', permission_classes=[permissions.IsAdminUser])
    def toggle_whatsapp(self, request):
        contenedor_id = request.data.get('id')
        if not contenedor_id:
            return Response({'mensaje': 'Faltan parametros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
        try:
            contenedor = Contenedor.objects.get(pk=contenedor_id)
            contenedor.acceso_whatsapp = not contenedor.acceso_whatsapp
            contenedor.save()
            return Response({
                'mensaje': f'WhatsApp {"activado" if contenedor.acceso_whatsapp else "desactivado"} para {contenedor.nombre}',
                'acceso_whatsapp': contenedor.acceso_whatsapp
            }, status=status.HTTP_200_OK)
        except Contenedor.DoesNotExist:
            return Response({'mensaje': 'El contenedor no existe', 'codigo': 15}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"], url_path=r'admin-lista', permission_classes=[permissions.IsAdminUser])
    def admin_lista(self, request):
        from contenedor.models import CtnWhatsappConexion
        contenedores = list(Contenedor.objects.exclude(schema_name='public').values(
            'id', 'schema_name', 'nombre', 'acceso_whatsapp', 'fecha', 'usuarios'
        ).order_by('nombre'))
        conexiones = {
            c.contenedor_id: c for c in CtnWhatsappConexion.objects.select_related('contenedor').all()
        }
        for c in contenedores:
            conexion = conexiones.get(c['id'])
            c['whatsapp_phone_number_id'] = conexion.phone_number_id if conexion else None
            c['whatsapp_display'] = conexion.display_phone_number if conexion else None
            c['whatsapp_estado'] = conexion.estado if conexion else None
        return Response(contenedores, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path=r'admin-whatsapp/numeros', permission_classes=[permissions.IsAdminUser])
    def admin_whatsapp_numeros(self, request):
        """Lista los numeros disponibles en la WABA de Rutenio + a quien estan asignados."""
        from mensajeria.servicios.admin_meta import AdminMetaServicio
        from contenedor.models import CtnWhatsappConexion

        servicio = AdminMetaServicio()
        resultado = servicio.listar_numeros()
        if resultado['error']:
            return Response({'mensaje': resultado['mensaje'], 'data': []}, status=status.HTTP_502_BAD_GATEWAY)

        asignaciones = {
            c.phone_number_id: {
                'contenedor_id': c.contenedor_id,
                'contenedor_nombre': c.contenedor.nombre,
                'schema_name': c.contenedor.schema_name,
                'estado': c.estado,
            }
            for c in CtnWhatsappConexion.objects.select_related('contenedor')
        }

        numeros = []
        for n in resultado['data']:
            asignado = asignaciones.get(n.get('id'))
            numeros.append({
                'phone_number_id': n.get('id'),
                'display_phone_number': n.get('display_phone_number'),
                'verified_name': n.get('verified_name'),
                'quality_rating': n.get('quality_rating'),
                'code_verification_status': n.get('code_verification_status'),
                'asignado_a': asignado,
            })
        return Response({'data': numeros}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path=r'admin-whatsapp/asignar', permission_classes=[permissions.IsAdminUser])
    def admin_whatsapp_asignar(self, request):
        """
        Asigna un numero a un contenedor. Body: {contenedor_id, phone_number_id}.
        Usa credenciales admin globales para poblar CtnWhatsappConexion.
        """
        from mensajeria.servicios.admin_meta import AdminMetaServicio
        from mensajeria.servicios.cifrado import CifradoServicio
        from contenedor.models import CtnWhatsappConexion
        import secrets

        contenedor_id = request.data.get('contenedor_id')
        phone_number_id = (request.data.get('phone_number_id') or '').strip()
        if not contenedor_id or not phone_number_id:
            return Response({'mensaje': 'contenedor_id y phone_number_id son requeridos', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)

        try:
            contenedor = Contenedor.objects.get(pk=contenedor_id)
        except Contenedor.DoesNotExist:
            return Response({'mensaje': 'Contenedor no existe', 'codigo': 15}, status=status.HTTP_404_NOT_FOUND)

        # Evitar asignar el mismo numero a 2 contenedores
        otro = CtnWhatsappConexion.objects.filter(phone_number_id=phone_number_id).exclude(contenedor=contenedor).first()
        if otro:
            return Response({
                'mensaje': f'Ese número ya está asignado a {otro.contenedor.nombre}',
                'codigo': 16,
            }, status=status.HTTP_400_BAD_REQUEST)

        servicio = AdminMetaServicio()
        detalle = servicio.consultar_numero(phone_number_id)
        if detalle['error']:
            return Response({'mensaje': f'No se pudo validar el número: {detalle["mensaje"]}'}, status=status.HTTP_502_BAD_GATEWAY)

        datos = detalle.get('data') or {}
        conexion, _ = CtnWhatsappConexion.objects.update_or_create(
            contenedor=contenedor,
            defaults={
                'phone_number_id': phone_number_id,
                'waba_id': servicio.waba_id,
                'display_phone_number': datos.get('display_phone_number'),
                'verified_name': datos.get('verified_name'),
                'access_token_cifrado': CifradoServicio.cifrar(servicio.access_token),
                'verify_token': secrets.token_urlsafe(24),
                'estado': CtnWhatsappConexion.ESTADO_ACTIVO,
                'error_mensaje': None,
            }
        )
        contenedor.acceso_whatsapp = True
        contenedor.save(update_fields=['acceso_whatsapp'])

        return Response({
            'mensaje': f'Número {datos.get("display_phone_number")} asignado a {contenedor.nombre}',
            'conexion': {
                'id': conexion.id,
                'phone_number_id': conexion.phone_number_id,
                'display_phone_number': conexion.display_phone_number,
                'estado': conexion.estado,
            }
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path=r'admin-whatsapp/desasignar', permission_classes=[permissions.IsAdminUser])
    def admin_whatsapp_desasignar(self, request):
        """Quita la asignación de WhatsApp a un contenedor. Body: {contenedor_id}."""
        from contenedor.models import CtnWhatsappConexion

        contenedor_id = request.data.get('contenedor_id')
        if not contenedor_id:
            return Response({'mensaje': 'contenedor_id requerido'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            contenedor = Contenedor.objects.get(pk=contenedor_id)
        except Contenedor.DoesNotExist:
            return Response({'mensaje': 'Contenedor no existe'}, status=status.HTTP_404_NOT_FOUND)

        CtnWhatsappConexion.objects.filter(contenedor=contenedor).delete()
        contenedor.acceso_whatsapp = False
        contenedor.save(update_fields=['acceso_whatsapp'])
        return Response({'mensaje': f'WhatsApp desasignado de {contenedor.nombre}'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path=r'admin-entregas', permission_classes=[permissions.IsAdminUser])
    def admin_entregas(self, request):
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        if not fecha_desde or not fecha_hasta:
            return Response(
                {'mensaje': 'Se requieren los parámetros fecha_desde y fecha_hasta', 'codigo': 1},
                status=status.HTTP_400_BAD_REQUEST
            )

        from ruteo.models.despacho import RutDespacho
        from ruteo.models.visita import RutVisita
        from ruteo.models.notificacion import RutNotificacion

        contenedores = Contenedor.objects.exclude(schema_name='public').values(
            'id', 'schema_name', 'nombre', 'fecha_ultima_conexion'
        ).order_by('nombre')

        resultados = []
        totales = {
            'total_despachos': 0,
            'visitas': 0,
            'visitas_entregadas': 0,
            'visitas_novedad': 0,
            'unidades': 0,
            'peso': 0.0,
            'volumen': 0.0,
            'decodificadas': 0,
            'whatsapp_enviados': 0,
        }

        for contenedor in contenedores:
            try:
                with schema_context(contenedor['schema_name']):
                    agregados = RutDespacho.objects.filter(
                        estado_aprobado=True,
                        estado_anulado=False,
                        fecha__date__gte=fecha_desde,
                        fecha__date__lte=fecha_hasta,
                    ).aggregate(
                        total_despachos=Count('id'),
                        visitas=Sum('visitas'),
                        visitas_entregadas=Sum('visitas_entregadas'),
                        visitas_novedad=Sum('visitas_novedad'),
                        unidades=Sum('unidades'),
                        peso=Sum('peso'),
                        volumen=Sum('volumen'),
                    )

                    decodificadas = RutVisita.objects.filter(
                        estado_decodificado=True,
                        fecha__date__gte=fecha_desde,
                        fecha__date__lte=fecha_hasta,
                    ).count()

                    whatsapp_enviados = RutNotificacion.objects.filter(
                        estado_enviado=True,
                        fecha__date__gte=fecha_desde,
                        fecha__date__lte=fecha_hasta,
                    ).count()

                datos = {
                    'contenedor_id': contenedor['id'],
                    'schema_name': contenedor['schema_name'],
                    'nombre': contenedor['nombre'] or contenedor['schema_name'],
                    'fecha_ultima_conexion': contenedor['fecha_ultima_conexion'],
                    'total_despachos': agregados['total_despachos'] or 0,
                    'visitas': agregados['visitas'] or 0,
                    'visitas_entregadas': agregados['visitas_entregadas'] or 0,
                    'visitas_novedad': agregados['visitas_novedad'] or 0,
                    'unidades': agregados['unidades'] or 0,
                    'peso': round(agregados['peso'] or 0, 1),
                    'volumen': round(agregados['volumen'] or 0, 1),
                    'decodificadas': decodificadas,
                    'whatsapp_enviados': whatsapp_enviados,
                }

                if datos['total_despachos'] > 0 or datos['decodificadas'] > 0 or datos['whatsapp_enviados'] > 0:
                    resultados.append(datos)
                    for key in totales:
                        totales[key] += datos[key]
            except Exception:
                continue

        totales['peso'] = round(totales['peso'], 1)
        totales['volumen'] = round(totales['volumen'], 1)

        formato = request.query_params.get('formato')
        if formato == 'xlsx':
            return self._generar_excel_entregas(resultados, totales, fecha_desde, fecha_hasta)

        return Response({
            'resultados': resultados,
            'totales': totales,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path=r'admin-entregas/(?P<schema_name>[^/.]+)', permission_classes=[permissions.IsAdminUser])
    def admin_entregas_detalle(self, request, schema_name=None):
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        if not fecha_desde or not fecha_hasta:
            return Response(
                {'mensaje': 'Se requieren los parámetros fecha_desde y fecha_hasta', 'codigo': 1},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            contenedor = Contenedor.objects.get(schema_name=schema_name)
        except Contenedor.DoesNotExist:
            return Response({'mensaje': 'Empresa no encontrada', 'codigo': 15}, status=status.HTTP_404_NOT_FOUND)

        from ruteo.models.despacho import RutDespacho
        from ruteo.models.visita import RutVisita

        with schema_context(schema_name):
            despachos = list(RutDespacho.objects.filter(
                estado_aprobado=True,
                estado_anulado=False,
                fecha__date__gte=fecha_desde,
                fecha__date__lte=fecha_hasta,
            ).select_related('vehiculo').order_by('-fecha').values(
                'id', 'fecha', 'fecha_salida', 'vehiculo__placa',
                'visitas', 'visitas_entregadas', 'visitas_novedad', 'visitas_liberadas',
                'unidades', 'peso', 'volumen',
                'estado_terminado', 'tiempo_servicio', 'tiempo_trayecto',
            ))

            despacho_ids = [d['id'] for d in despachos]

            visitas = list(RutVisita.objects.filter(
                despacho_id__in=despacho_ids,
            ).order_by('despacho_id', 'orden').values(
                'id', 'numero', 'documento', 'destinatario',
                'destinatario_direccion', 'destinatario_telefono',
                'unidades', 'peso',
                'estado_entregado', 'estado_novedad', 'estado_decodificado',
                'orden', 'despacho_id',
                'datos_entrega',
            ))

        visitas_por_despacho = {}
        for v in visitas:
            did = v.pop('despacho_id')
            visitas_por_despacho.setdefault(did, []).append(v)

        for d in despachos:
            d['visitas_detalle'] = visitas_por_despacho.get(d['id'], [])

        return Response({
            'empresa': {
                'contenedor_id': contenedor.id,
                'schema_name': contenedor.schema_name,
                'nombre': contenedor.nombre or contenedor.schema_name,
            },
            'despachos': despachos,
        }, status=status.HTTP_200_OK)

    def _generar_excel_entregas(self, resultados, totales, fecha_desde, fecha_hasta):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = 'Entregas por empresa'

        encabezados = [
            'Empresa', 'Subdominio', 'Decodificadas', 'WhatsApp', 'Despachos', 'Visitas',
            'Entregadas', 'Novedades', 'Unidades', 'Peso (kg)', 'Volumen (m³)'
        ]
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0098D7', end_color='0098D7', fill_type='solid')

        for col, titulo in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col, value=titulo)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = Alignment(horizontal='center')

        for fila, empresa in enumerate(resultados, 2):
            ws.cell(row=fila, column=1, value=empresa['nombre'])
            ws.cell(row=fila, column=2, value=empresa['schema_name'])
            ws.cell(row=fila, column=3, value=empresa['decodificadas'])
            ws.cell(row=fila, column=4, value=empresa['whatsapp_enviados'])
            ws.cell(row=fila, column=5, value=empresa['total_despachos'])
            ws.cell(row=fila, column=6, value=empresa['visitas'])
            ws.cell(row=fila, column=7, value=empresa['visitas_entregadas'])
            ws.cell(row=fila, column=8, value=empresa['visitas_novedad'])
            ws.cell(row=fila, column=9, value=empresa['unidades'])
            ws.cell(row=fila, column=10, value=empresa['peso'])
            ws.cell(row=fila, column=11, value=empresa['volumen'])

        fila_total = len(resultados) + 2
        total_font = Font(bold=True)
        ws.cell(row=fila_total, column=1, value='TOTAL').font = total_font
        ws.cell(row=fila_total, column=3, value=totales['decodificadas']).font = total_font
        ws.cell(row=fila_total, column=4, value=totales['whatsapp_enviados']).font = total_font
        ws.cell(row=fila_total, column=5, value=totales['total_despachos']).font = total_font
        ws.cell(row=fila_total, column=6, value=totales['visitas']).font = total_font
        ws.cell(row=fila_total, column=7, value=totales['visitas_entregadas']).font = total_font
        ws.cell(row=fila_total, column=8, value=totales['visitas_novedad']).font = total_font
        ws.cell(row=fila_total, column=9, value=totales['unidades']).font = total_font
        ws.cell(row=fila_total, column=10, value=totales['peso']).font = total_font
        ws.cell(row=fila_total, column=11, value=totales['volumen']).font = total_font

        for col in range(1, 12):
            ws.column_dimensions[chr(64 + col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="entregas_{fecha_desde}_{fecha_hasta}.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path=r'conectar',)
    def conectar(self, request):
        raw = request.data
        subdominio = raw.get('subdominio', None)
        if subdominio:
            try:
                contenedor = Contenedor.objects.get(schema_name=subdominio)
                contenedor.fecha_ultima_conexion = timezone.now()
                contenedor.save()
                contenedor_serializador = ContenedorSerializador(contenedor)
                return Response(contenedor_serializador.data, status=status.HTTP_200_OK)
            except Contenedor.DoesNotExist:
                return Response({'mensaje':'El contenedor no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)  

                    
