"""Plantilla Excel corporativa reutilizable para exportaciones.

Objetivo: que todas las exportaciones compartan una misma cara (encabezado con
logo + empresa + titulo + rango + fecha de generacion, columnas estilizadas,
auto-ancho, inmovilizar encabezado, autofiltro, cebra y fila de totales) en vez
de que cada una arme su estilo por su cuenta.

Uso:
    plantilla = ExcelPlantilla('Entregas por zona', 'Del 2026-07-01 al 2026-07-31')
    plantilla.agregar_hoja(
        'Resumen',
        columnas=[
            {'clave': 'mensajero', 'titulo': 'Mensajero'},
            {'clave': 'asignadas', 'titulo': 'Asignadas', 'tipo': 'entero'},
        ],
        filas=[{'mensajero': 'Ana', 'asignadas': 12}, ...],
        totales=['asignadas'],
    )
    return plantilla.respuesta('entregas_por_zona.xlsx')
"""
from datetime import datetime, date
from io import BytesIO

import requests
from decouple import config
from django.http import HttpResponse
from django.utils import timezone
from django.utils.encoding import smart_str
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from general.models.empresa import GenEmpresa


class ExcelPlantilla:
    COLOR_BANDA = '1F3B57'       # azul oscuro del encabezado de marca
    COLOR_ENCABEZADO = '2F5F8A'  # fila de titulos de columna
    COLOR_CEBRA = 'F2F6FA'       # filas pares
    COLOR_TOTAL = 'DDE7F1'       # fila de totales
    ANCHO_MAX = 48

    FORMATOS = {
        'numero': '#,##0.00',
        'entero': '#,##0',
        'fecha': 'yyyy-mm-dd',
        'fecha_hora': 'yyyy-mm-dd hh:mm',
    }

    def __init__(self, titulo, subtitulo=None):
        self.titulo = titulo
        self.subtitulo = subtitulo
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # empezar sin hoja por defecto
        self._empresa = GenEmpresa.objects.filter(pk=1).first()
        self._logo_bytes = None  # None = no intentado; b'' = intentado sin exito
        self._borde = Border(*[Side(style='thin', color='D9D9D9')] * 4)

    # -- logo ---------------------------------------------------------------
    def _obtener_logo_bytes(self):
        if self._logo_bytes is not None:
            return self._logo_bytes or None
        self._logo_bytes = b''
        try:
            if self._empresa and self._empresa.imagen:
                region = config('DO_REGION')
                bucket = config('DO_BUCKET')
                url = f'https://{bucket}.{region}.digitaloceanspaces.com/{self._empresa.imagen}'
                resp = requests.get(url, timeout=5)
                if resp.status_code == 200 and resp.content:
                    self._logo_bytes = resp.content
        except Exception:
            self._logo_bytes = b''
        return self._logo_bytes or None

    def _imagen_logo(self, alto=52):
        datos = self._obtener_logo_bytes()
        if not datos:
            return None
        try:
            from PIL import Image as PILImage
            with PILImage.open(BytesIO(datos)) as im:
                ancho, altura = im.size
            img = XLImage(BytesIO(datos))
            img.height = alto
            img.width = int(ancho * (alto / altura)) if altura else alto
            return img
        except Exception:
            return None

    # -- valores ------------------------------------------------------------
    @staticmethod
    def _sin_tz(valor):
        if isinstance(valor, datetime) and valor.tzinfo is not None:
            return timezone.localtime(valor).replace(tzinfo=None)
        return valor

    def _valor(self, valor, tipo):
        if valor is None:
            return ''
        if tipo == 'bool':
            return 'SI' if valor else 'NO'
        if tipo in ('fecha', 'fecha_hora'):
            return self._sin_tz(valor)
        if tipo in ('numero', 'entero'):
            return valor
        limpio = ILLEGAL_CHARACTERS_RE.sub('', smart_str(valor))
        return limpio

    # -- hoja ---------------------------------------------------------------
    def agregar_hoja(self, nombre, columnas, filas, totales=None):
        """columnas: [{'clave','titulo','tipo'?,'ancho'?}] | filas: [dict] | totales: [clave]."""
        totales = set(totales or [])
        ws = self.wb.create_sheet(title=nombre[:31])
        ncols = len(columnas)
        ultima = get_column_letter(ncols)

        self._pintar_banda(ws, ncols, ultima)

        fila_encabezado = 6
        fila_datos = fila_encabezado + 1

        # Encabezados de columna.
        fuente_enc = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        relleno_enc = PatternFill('solid', fgColor=self.COLOR_ENCABEZADO)
        for idx, col in enumerate(columnas, start=1):
            celda = ws.cell(row=fila_encabezado, column=idx, value=col['titulo'])
            celda.font = fuente_enc
            celda.fill = relleno_enc
            celda.alignment = Alignment(horizontal='center', vertical='center')
            celda.border = self._borde

        # Datos.
        fuente = Font(name='Arial', size=10)
        relleno_cebra = PatternFill('solid', fgColor=self.COLOR_CEBRA)
        for i, fila in enumerate(filas):
            r = fila_datos + i
            cebra = i % 2 == 1
            for idx, col in enumerate(columnas, start=1):
                tipo = col.get('tipo', 'texto')
                celda = ws.cell(row=r, column=idx, value=self._valor(fila.get(col['clave']), tipo))
                celda.font = fuente
                celda.border = self._borde
                celda.alignment = Alignment(
                    horizontal='right' if tipo in ('numero', 'entero') else 'left',
                    vertical='center',
                )
                if tipo in self.FORMATOS:
                    celda.number_format = self.FORMATOS[tipo]
                if cebra:
                    celda.fill = relleno_cebra

        # Fila de totales.
        if totales and filas:
            r = fila_datos + len(filas)
            fuente_total = Font(name='Arial', size=10, bold=True)
            relleno_total = PatternFill('solid', fgColor=self.COLOR_TOTAL)
            borde_total = Border(top=Side(style='medium', color='9DB4CC'))
            for idx, col in enumerate(columnas, start=1):
                celda = ws.cell(row=r, column=idx)
                celda.font = fuente_total
                celda.fill = relleno_total
                celda.border = borde_total
                if idx == 1:
                    celda.value = 'TOTAL'
                elif col['clave'] in totales:
                    celda.value = sum(
                        f.get(col['clave']) or 0
                        for f in filas
                        if isinstance(f.get(col['clave']), (int, float))
                    )
                    celda.number_format = self.FORMATOS.get(col.get('tipo'), '#,##0')
                    celda.alignment = Alignment(horizontal='right')

        self._auto_ancho(ws, columnas, filas)
        ws.freeze_panes = ws.cell(row=fila_datos, column=1)
        if filas:
            ws.auto_filter.ref = f'A{fila_encabezado}:{ultima}{fila_datos + len(filas) - 1}'
        return ws

    def _pintar_banda(self, ws, ncols, ultima):
        relleno = PatternFill('solid', fgColor=self.COLOR_BANDA)
        empresa = self._empresa.nombre_corto if self._empresa else ''
        nit = (self._empresa.numero_identificacion or '') if self._empresa else ''
        linea_empresa = empresa + (f'  ·  NIT: {nit}' if nit else '')
        generado = f"Generado: {self._sin_tz(timezone.now()).strftime('%Y-%m-%d %H:%M')}"

        textos = [
            (self.titulo, Font(name='Arial', size=14, bold=True, color='FFFFFF'), 'left'),
            (linea_empresa, Font(name='Arial', size=10, bold=True, color='FFFFFF'), 'left'),
            (self.subtitulo or '', Font(name='Arial', size=9, italic=True, color='D6E2EF'), 'left'),
            (generado, Font(name='Arial', size=8, color='D6E2EF'), 'right'),
        ]
        for i, (texto, fuente, alineacion) in enumerate(textos, start=1):
            # El relleno se pinta ANTES de fusionar: tras merge_cells las celdas
            # no ancla quedan de solo lectura y asignarles estilo revienta.
            for c in range(1, ncols + 1):
                ws.cell(row=i, column=c).fill = relleno
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ncols)
            celda = ws.cell(row=i, column=1)
            celda.value = texto
            celda.font = fuente
            celda.alignment = Alignment(horizontal=alineacion, vertical='center', indent=1)
            ws.row_dimensions[i].height = 20 if i == 1 else 15

        logo = self._imagen_logo()
        if logo is not None:
            ws.add_image(logo, f'{get_column_letter(max(1, ncols))}1')

    def _auto_ancho(self, ws, columnas, filas):
        for idx, col in enumerate(columnas, start=1):
            ancho = col.get('ancho')
            if not ancho:
                largo = len(str(col['titulo']))
                for fila in filas:
                    valor = fila.get(col['clave'])
                    if valor is not None:
                        largo = max(largo, len(str(valor)))
                ancho = min(largo + 3, self.ANCHO_MAX)
            ws.column_dimensions[get_column_letter(idx)].width = ancho

    # -- salida -------------------------------------------------------------
    def respuesta(self, nombre_archivo):
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response
