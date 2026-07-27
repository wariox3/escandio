"""Reporte de entregas por zona (factor de pago del mensajero).

GET /ruteo/reporte/mensajero/entregas/ devuelve:
  - 'resumen': conteo por (mensajero x zona). Alimenta el pago -> NUNCA se
    trunca (se agrega en BD, aparte del detalle).
  - 'relacion': detalle guia por guia, acotado a LIMITE_RELACION.

La zona vive denormalizada en la visita (franja_id/franja_codigo); el nombre se
resuelve contra RutFranja. Una guia sin franja sale con zona en null (no se
descarta), porque para pago hay que verla y corregirla.

Correr: python manage.py test ruteo.tests_reporte_entregas
"""
from django_tenants.test.cases import TenantTestCase

from contenedor.models import User
from ruteo.models.despacho import RutDespacho
from ruteo.models.franja import RutFranja
from ruteo.models.vehiculo import RutVehiculo
from ruteo.models.visita import RutVisita
from ruteo.views.reporte import ReporteMensajeroEntregasView


class _Req:
    """Solo se usa request.query_params.get(...)."""

    def __init__(self, params=None):
        self.query_params = params or {}


class ReporteEntregasZonaTests(TenantTestCase):

    def setUp(self):
        super().setUp()
        self.mensajero = User.objects.create(
            username='m@x.com', correo='m@x.com', nombre='Ana', apellido='Ruiz', is_active=True,
        )
        self.z1 = RutFranja.objects.create(codigo='Z1', nombre='Norte')
        self.z2 = RutFranja.objects.create(codigo='Z2', nombre='Sur')
        self.despacho = RutDespacho.objects.create(conductor_id=self.mensajero.id)

    def _visita(self, franja=None, entregado=False, novedad=False, despacho=None):
        return RutVisita.objects.create(
            despacho=despacho or self.despacho,
            ciudad_id=None,
            estado_despacho=True,
            franja_id=franja.id if franja else None,
            franja_codigo=franja.codigo if franja else None,
            estado_entregado=entregado,
            estado_novedad=novedad,
        )

    def _get(self, params=None, limite=None):
        vista = ReporteMensajeroEntregasView()
        if limite is not None:
            vista.LIMITE_RELACION = limite
        return vista.get(_Req(params)).data

    def test_resumen_agrupa_por_mensajero_y_zona_con_conteos(self):
        self._visita(self.z1, entregado=True)
        self._visita(self.z1, novedad=True)
        self._visita(self.z2, entregado=True)

        resumen = {r['zona_codigo']: r for r in self._get()['resumen']}

        self.assertEqual(resumen['Z1']['asignadas'], 2)
        self.assertEqual(resumen['Z1']['entregadas'], 1)
        self.assertEqual(resumen['Z1']['novedades'], 1)
        self.assertEqual(resumen['Z1']['conductor_nombre'], 'Ana Ruiz')
        self.assertEqual(resumen['Z1']['zona_nombre'], 'Norte')
        self.assertEqual(resumen['Z2']['asignadas'], 1)
        self.assertEqual(resumen['Z2']['entregadas'], 1)

    def test_guia_sin_zona_no_se_descarta(self):
        self._visita(franja=None, entregado=True)

        data = self._get()
        self.assertEqual(len(data['relacion']), 1)
        self.assertIsNone(data['relacion'][0]['zona_id'])
        self.assertIsNone(data['relacion'][0]['zona_nombre'])
        # Y tambien cuenta en el resumen, con zona nula.
        self.assertEqual(len(data['resumen']), 1)
        self.assertIsNone(data['resumen'][0]['zona_id'])
        self.assertEqual(data['resumen'][0]['asignadas'], 1)

    def test_estado_se_deriva(self):
        self._visita(self.z1, entregado=True)
        self._visita(self.z1, novedad=True)
        self._visita(self.z1)

        estados = sorted(r['estado'] for r in self._get()['relacion'])
        self.assertEqual(estados, ['entregada', 'novedad', 'pendiente'])

    def test_relacion_trae_zona_nombre_y_mensajero(self):
        self._visita(self.z1, entregado=True)
        fila = self._get()['relacion'][0]
        self.assertEqual(fila['zona_nombre'], 'Norte')
        self.assertEqual(fila['zona_codigo'], 'Z1')
        self.assertEqual(fila['conductor_nombre'], 'Ana Ruiz')

    def test_sin_mensajero_pero_con_placa_conserva_la_placa(self):
        """Despacho sin conductor pero con vehiculo: la placa no se pierde en el
        resumen, para poder atribuir el pago o corregir la asignacion."""
        vehiculo = RutVehiculo.objects.create(placa='XYZ789')
        despacho = RutDespacho.objects.create(conductor_id=None, vehiculo=vehiculo)
        self._visita(self.z1, entregado=True, despacho=despacho)

        resumen = self._get()['resumen']
        self.assertEqual(len(resumen), 1)
        self.assertIsNone(resumen[0]['conductor_id'])
        self.assertEqual(resumen[0]['placa'], 'XYZ789')
        self.assertEqual(resumen[0]['asignadas'], 1)

    def test_despacho_anulado_se_excluye(self):
        anulado = RutDespacho.objects.create(conductor_id=self.mensajero.id, estado_anulado=True)
        self._visita(self.z1, entregado=True, despacho=anulado)

        data = self._get()
        self.assertEqual(len(data['relacion']), 0)
        self.assertEqual(len(data['resumen']), 0)

    def test_exportacion_excel_se_genera_y_abre(self):
        """El export a Excel responde un xlsx valido con las dos hojas."""
        from io import BytesIO
        from openpyxl import load_workbook

        self._visita(self.z1, entregado=True)
        respuesta = ReporteMensajeroEntregasView().get(_Req({'excel': '1'}))

        self.assertEqual(respuesta.status_code, 200)
        self.assertIn('spreadsheetml', respuesta['Content-Type'])
        self.assertIn('attachment', respuesta['Content-Disposition'])
        wb = load_workbook(BytesIO(respuesta.content))
        self.assertEqual(wb.sheetnames, ['Resumen por zona', 'Relación'])

        hoja = wb['Resumen por zona']
        self.assertIn('Entregas por zona', str(hoja['A1'].value))  # banda de titulo
        self.assertEqual(hoja['A6'].value, 'Mensajero')            # encabezado de columna
        self.assertEqual(hoja['A7'].value, 'Ana Ruiz')             # dato
        # Fila de totales al final con la suma de asignadas.
        self.assertEqual(hoja['A8'].value, 'TOTAL')
        self.assertEqual(hoja['E8'].value, 1)

    def test_resumen_no_se_trunca_aunque_la_relacion_si(self):
        """Propiedad de pago: el detalle se acota, los totales van completos."""
        for _ in range(5):
            self._visita(self.z1, entregado=True)

        data = self._get(limite=2)

        self.assertTrue(data['truncado'])
        self.assertEqual(data['relacion_count'], 2)
        # El resumen (pago) suma las 5, no las 2 del detalle acotado.
        self.assertEqual(data['resumen'][0]['asignadas'], 5)
        self.assertEqual(data['resumen'][0]['entregadas'], 5)
