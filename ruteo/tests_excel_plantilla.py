"""ExcelPlantilla: comportamiento del helper que migra listas a la plantilla.

Correr: python manage.py test ruteo.tests_excel_plantilla
"""
from io import BytesIO

from django_tenants.test.cases import TenantTestCase
from openpyxl import load_workbook

from utilidades.excel_plantilla import ExcelPlantilla


class ExcelPlantillaTests(TenantTestCase):

    def _hoja(self, plantilla, nombre):
        wb = load_workbook(BytesIO(plantilla.respuesta('x.xlsx').content))
        return wb[nombre]

    def test_decimal_texto_se_escribe_y_suma_como_numero(self):
        """Los decimales que DRF manda como texto deben quedar numericos."""
        plantilla = ExcelPlantilla('T')
        plantilla.agregar_hoja_datos(
            'H',
            [{'zona': 'A', 'peso': '12.50'}, {'zona': 'B', 'peso': '7.5'}],
            tipos={'peso': 'numero'},
            totales=['peso'],
        )
        hoja = self._hoja(plantilla, 'H')
        self.assertEqual(hoja['A6'].value, 'Zona')   # encabezado (fila 6)
        self.assertEqual(hoja['B7'].value, 12.5)     # numero, no texto
        self.assertIsInstance(hoja['B7'].value, float)
        self.assertEqual(hoja['B9'].value, 20.0)     # total sumado numericamente

    def test_bool_como_si_no_y_titulo_legible(self):
        plantilla = ExcelPlantilla('T')
        plantilla.agregar_hoja_datos(
            'H',
            [{'estado_entregado': True, 'despacho__vehiculo__placa': 'ABC'}],
            titulos={'despacho__vehiculo__placa': 'Placa'},
            tipos={'estado_entregado': 'bool'},
        )
        hoja = self._hoja(plantilla, 'H')
        self.assertEqual(hoja['A6'].value, 'Estado entregado')  # fallback legible
        self.assertEqual(hoja['B6'].value, 'Placa')             # titulo explicito
        self.assertEqual(hoja['A7'].value, 'SI')

    def test_logo_ruteo_va_incrustado(self):
        """El logo de Ruteo (asset local) debe quedar en la banda."""
        plantilla = ExcelPlantilla('T')
        plantilla.agregar_hoja_datos('H', [{'a': 1}])
        # Sin empresa/red no hay logo de cliente, pero el de Ruteo (local) si.
        self.assertGreaterEqual(len(plantilla.wb['H']._images), 1)

    def test_datos_vacios_no_revienta(self):
        plantilla = ExcelPlantilla('T')
        plantilla.agregar_hoja_datos('H', [])
        wb = load_workbook(BytesIO(plantilla.respuesta('x.xlsx').content))
        self.assertIn('H', wb.sheetnames)
