"""Export a Excel del reporte por mensajero (conteos) con la plantilla.

GET /ruteo/reporte/mensajero/?excel=1 agrega por mensajero/placa/dia en el
backend (misma cuenta que hace el front en pantalla) y devuelve un xlsx con
tres hojas: Detalle diario, Totales por mensajero y Totales por placa.

Correr: python manage.py test ruteo.tests_reporte_mensajero_excel
"""
from io import BytesIO

from django_tenants.test.cases import TenantTestCase
from openpyxl import load_workbook

from ruteo.views.reporte import ReporteMensajeroView


class _Req:
    def __init__(self, params=None):
        self.query_params = params or {}


class ReporteMensajeroExcelTests(TenantTestCase):

    def _exportar(self, resultados):
        respuesta = ReporteMensajeroView()._exportar_excel(resultados, None, None)
        return load_workbook(BytesIO(respuesta.content))

    def test_tres_hojas_y_totales(self):
        # Dos despachos del mismo mensajero/placa/dia -> una fila agregada.
        resultados = [
            {'id': 1, 'fecha': None, 'conductor_id': 7, 'conductor_nombre': 'Ana',
             'vehiculo__placa': 'ABC', 'visitas': 10, 'visitas_entregadas': 8, 'visitas_novedad': 2},
            {'id': 2, 'fecha': None, 'conductor_id': 7, 'conductor_nombre': 'Ana',
             'vehiculo__placa': 'ABC', 'visitas': 5, 'visitas_entregadas': 5, 'visitas_novedad': 0},
        ]
        wb = self._exportar(resultados)
        self.assertEqual(
            wb.sheetnames,
            ['Detalle diario', 'Totales por mensajero', 'Totales por placa'],
        )
        hoja = wb['Detalle diario']
        self.assertEqual(hoja['A6'].value, 'Mensajero')
        self.assertEqual(hoja['A7'].value, 'Ana')
        # despachos=2, asignadas=15, entregadas=13 en la fila agregada.
        self.assertEqual(hoja['D7'].value, 2)
        self.assertEqual(hoja['E7'].value, 15)
        self.assertEqual(hoja['F7'].value, 13)
        # Fila de totales suma asignadas.
        self.assertEqual(hoja['A8'].value, 'TOTAL')
        self.assertEqual(hoja['E8'].value, 15)

    def test_sin_asignar_y_sin_placa(self):
        resultados = [
            {'id': 1, 'fecha': None, 'conductor_id': None, 'conductor_nombre': None,
             'vehiculo__placa': None, 'visitas': 3, 'visitas_entregadas': 1, 'visitas_novedad': 1},
        ]
        hoja = self._exportar(resultados)['Detalle diario']
        self.assertEqual(hoja['A7'].value, 'Sin asignar')
        self.assertEqual(hoja['B7'].value, 'Sin placa')
